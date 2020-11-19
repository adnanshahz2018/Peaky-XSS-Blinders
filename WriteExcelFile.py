
import os
import xlwt
import xlsxwriter 
import openpyxl as op
from xlwt import Workbook

class write_excel_file:
    filename = ''
    payload = ''
    url = ''

    def __init__(self, filename, payload):
        self.filename = filename
        self.payload = payload

    def create_attack_storage_excel_file(self, filepath):
        # creating the file for the first time 
        if not os.path.exists(filepath):
            workbook = xlsxwriter.Workbook(filepath)
            worksheet = workbook.add_worksheet("Sheet1")
            workbook.close()
            wb = op.load_workbook(filepath, False)
            ws = wb['Sheet1']
            ws.append(["Attack URL", "Context", "Success", "Detection" ])
            wb.save(filepath)
            wb.close()


    def write_contexts(self, url, attrs, htmls, scripts, urls):
        # creating the file for the first time 
        if not os.path.exists(self.filename):
            workbook = xlsxwriter.Workbook(self.filename)
            worksheet = workbook.add_worksheet("data")
            workbook.close()
            wb = op.load_workbook(self.filename, False)
            ws = wb['data']
            ws.append(["web-url", "payload", "# attr ", "# html ", "# script", "# url ",
                        "attr ", "html ", "script ", "url "    ])
            wb.save(self.filename)
            wb.close()

        print('\n\t\t --------- Writing to EXCEL ------------')
        wb = op.load_workbook(self.filename, False)
        ws = wb['data']
        ws.append([ str(url),   self.payload,
                    len(attrs), len(htmls), len(scripts), len(urls),
                    '',     '',     '',     '' ] )

        if not attrs == 'None':
            for attr in attrs:
                ws.cell(row=ws.max_row, column=7).value += str(attr) + ' , '
        else:
            ws.cell(row=ws.max_row, column=7).value += str(attrs)

        if not htmls == 'None':
            for html in htmls:  
                ws.cell(row=ws.max_row, column=8).value += str(html) + ' , '
        else:
            ws.cell(row=ws.max_row, column=8).value += str(htmls)

        if not scripts == 'None':
            for script in scripts:
                ws.cell(row=ws.max_row, column=9).value += str(script) + ' , '
        else:
            ws.cell(row=ws.max_row, column=9).value += str(scripts)

        if not urls == 'None':
            for url in urls:
                ws.cell(row=ws.max_row, column=10).value += str(url) + ' , '
        else:
            ws.cell(row=ws.max_row, column=10).value += str(urls)
            
        wb.save(self.filename)
        wb.close()

if __name__ == "__main__":
	print('{WriteExcelFile}')